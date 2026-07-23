from openpyxl import load_workbook

names = ["John", "Mary", "Alex", "Sophia"]
salaries = [60000, 75000, 50000, 80000]

new_list =[]
for name,sal in zip(names,salaries):
    print(name,sal)    
    new_list.append({
        "name" : name,
        "salary": sal
    })

sorted_list = sorted(new_list, key=lambda emp: emp["salary"],reverse=True)
print(sorted_list)

for i,li in enumerate(sorted_list,start=1):
    print(f"Rank {i}: {li['name']} - {li['salary']}")


avg = lambda a , b ,c : (a+b+c)/3
print(avg(3,6,9))

workbook = load_workbook("employees.xlsx")
sheet = workbook.active

row1 = sheet["A2"]
#print(row1.value)
row2 = sheet.cell(row=2,column=1).value
##print(row2)

print(sheet.dimensions)
print(sheet["A"])
print(sheet["A:A"])
print(sheet.rows)
print(sheet.columns)