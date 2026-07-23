from openpyxl import load_workbook

workbook = load_workbook("employees.xlsx")
sheet = workbook.active

sheet["F1"] = "Bonus"
for i in range(2, sheet.max_row + 1):
    salary = sheet.cell(row=i, column=4).value
    if salary >=  70000:
        sheet.cell(row=i, column=6).value = 10000
    elif salary >= 50000:
        sheet.cell(row=i, column=6).value = 5000
    else:
        sheet.cell(row=i, column=6).value = 2000

sheet["G1"] = "Grade"
for i in range(2, sheet.max_row +1):
    Salary = sheet.cell(row=i,column=4).value
    if Salary >= 80000:
        sheet.cell(row=i, column=7).value = "A"
    elif Salary >= 70000:
        sheet.cell(row=i, column=7).value = "B"
    elif Salary >= 60000:
        sheet.cell(row=i, column=7).value = "C"
    elif Salary >= 50000:
        sheet.cell(row=i, column=7).value = "D"
    else:
        sheet.cell(row=i, column=7).value = "E"

hikes = {
    "IT": 15,
    "Finance": 12,
    "HR": 10,
    "Sales": 8
}
sheet["H1"] = "Hike %"
sheet["I1"] = "New Salary"

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row):
    department = row[2].value
    salary = row[3].value

    hike = hikes[department]
    new_salary = salary + salary * hike / 100

    row[7].value = hike
    row[8].value = new_salary

workbook.save("employees.xlsx")

print("Total Employees Updated : ",(sheet.max_row-1))

highest_sal = sheet.cell(row=2,column=9).value
lowest_sal = highest_sal
total_sal = 0

for i in range(2,sheet.max_row +1):
    salary = sheet.cell(row=i,column=9).value
    total_sal += salary

    if salary > highest_sal:
        highest_sal = salary

    if salary <lowest_sal:
        lowest_sal = salary

print("Highest New Salary : ",highest_sal)
print("Lowest New Salary : ",lowest_sal)
print("Average New Salary : ",(total_sal / (sheet.max_row-1)))